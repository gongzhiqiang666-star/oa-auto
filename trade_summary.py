#!/usr/bin/env python3
from __future__ import annotations

import argparse
from collections import defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SOURCE_SHEET = "Sheet1"
SUMMARY_SHEET = "Sheet2"
NOTES_SHEET = "问题备注"
ALGORITHM_SHEET = "算法说明"

REQUIRED_HEADERS = {
    "所属二级",
    "所属通道",
    "商户费率",
    "支付类型",
    "交易金额",
    "交易状态",
}

CHANNEL_KEYWORDS = [
    ("联动", "力pos交易"),
    ("鲲鹏", "鲲鹏交易"),
    ("合利宝", "合利宝"),
    ("星驿付", "星驿付"),
]


def money(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def money_float(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def rate_header(rate) -> str:
    try:
        normalized = Decimal(str(rate)).normalize()
    except (InvalidOperation, TypeError):
        return f"{rate}交易额"
    return f"{normalized}交易额"


def channel_group(channel) -> str | None:
    if not channel:
        return None
    text = str(channel)
    for keyword, group in CHANNEL_KEYWORDS:
        if keyword in text:
            return group
    return text


def effective_header(pay_type, rate) -> str:
    if pay_type and str(pay_type).startswith("特惠"):
        return str(pay_type)
    return rate_header(rate)


def merged_title(ws, col: int) -> str | None:
    value = ws.cell(1, col).value
    if value:
        return str(value)
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= 1 <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            merged_value = ws.cell(cell_range.min_row, cell_range.min_col).value
            return str(merged_value) if merged_value else None
    return None


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def set_summary_rows(ws, people: list[str]) -> None:
    start_row = 4
    template_row = start_row
    current_rows = max(ws.max_row - start_row + 1, 0)
    needed_rows = len(people)

    if needed_rows > current_rows:
        ws.insert_rows(start_row + current_rows, needed_rows - current_rows)
    elif current_rows > needed_rows:
        ws.delete_rows(start_row + needed_rows, current_rows - needed_rows)

    for offset, person in enumerate(people):
        row = start_row + offset
        copy_row_style(ws, template_row, row)
        ws.cell(row, 1).value = person
        for col in range(2, ws.max_column + 1):
            ws.cell(row, col).value = 0


def clear_notes_sheet(wb):
    if NOTES_SHEET in wb.sheetnames:
        del wb[NOTES_SHEET]
    ws = wb.create_sheet(NOTES_SHEET)
    headers = ["类型", "说明", "姓名", "通道", "分类", "金额", "原始位置/数量"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    widths = [16, 48, 16, 16, 18, 14, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    return ws


def clear_algorithm_sheet(wb):
    if ALGORITHM_SHEET in wb.sheetnames:
        del wb[ALGORITHM_SHEET]
    ws = wb.create_sheet(ALGORITHM_SHEET)
    rows = [
        ("步骤", "说明"),
        ("1", "读取 Sheet1，每一行作为一笔交易明细。"),
        ("2", "只统计交易状态为“交易成功”，且有姓名、通道、费率的明细。"),
        ("3", "姓名取 Sheet1 的“所属二级”，金额取“交易金额”。"),
        ("4", "通道按关键字归类：联动=力pos交易，鲲鹏=鲲鹏交易，合利宝=合利宝，星驿付=星驿付。"),
        ("5", "普通交易按“商户费率”汇总，例如 0.6000 写入 0.6交易额。"),
        ("6", "支付类型以“特惠”开头的交易，按支付类型汇总到对应特惠列。"),
        ("7", "同一姓名、同一通道、同一费率/特惠类型的交易金额相加后写入 Sheet2。"),
        ("8", "每个通道的“总交易额”统计该姓名在该通道下的全部交易金额，即使模板没有独立费率列也计入总额。"),
        ("9", "模板缺列、人员不匹配、无法识别的手工合计行等问题，写入“问题备注”。"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 110
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def append_note(ws, kind, message, person="", channel="", category="", amount=None, location=""):
    amount_value = money_float(amount) if isinstance(amount, Decimal) else amount
    ws.append([kind, message, person, channel, category, amount_value, location])


def build_summary(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    if SOURCE_SHEET not in wb.sheetnames or SUMMARY_SHEET not in wb.sheetnames:
        raise RuntimeError(f"工作簿必须包含 {SOURCE_SHEET} 和 {SUMMARY_SHEET}")

    detail_ws = wb[SOURCE_SHEET]
    summary_ws = wb[SUMMARY_SHEET]
    notes_ws = clear_notes_sheet(wb)
    clear_algorithm_sheet(wb)

    headers = [cell.value for cell in detail_ws[1]]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise RuntimeError("Sheet1 缺少必要表头: " + ", ".join(sorted(missing)))
    idx = {header: pos for pos, header in enumerate(headers)}

    summary_columns = {}
    total_columns = {}
    for col in range(2, summary_ws.max_column + 1):
        group = merged_title(summary_ws, col)
        header = summary_ws.cell(3, col).value
        if not group or not header:
            continue
        if header == "总交易额":
            total_columns[group] = col
        else:
            summary_columns[(group, str(header))] = col

    template_people = []
    for row in range(4, summary_ws.max_row + 1):
        person = summary_ws.cell(row, 1).value
        if person:
            template_people.append(str(person))

    sums = defaultdict(Decimal)
    totals = defaultdict(Decimal)
    unsupported = defaultdict(Decimal)
    people_seen = set()
    skipped_rows = []
    detail_count = 0

    for row_no, row in enumerate(detail_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        person = row[idx["所属二级"]]
        channel = row[idx["所属通道"]]
        rate = row[idx["商户费率"]]
        pay_type = row[idx["支付类型"]]
        amount = money(row[idx["交易金额"]])
        status = row[idx["交易状态"]]

        if not person or not channel or not rate:
            skipped_rows.append((row_no, amount))
            continue
        if status != "交易成功":
            append_note(notes_ws, "跳过", "交易状态不是交易成功，未参与汇总", person, channel, "", amount, f"Sheet1!{row_no}")
            continue

        detail_count += 1
        person = str(person)
        group = channel_group(channel)
        category = effective_header(pay_type, rate)
        people_seen.add(person)
        totals[(person, group)] += amount

        if (group, category) in summary_columns:
            sums[(person, group, category)] += amount
        else:
            unsupported[(person, group, category)] += amount

    people = []
    for person in template_people:
        if person not in people:
            people.append(person)
    for person in sorted(people_seen):
        if person not in people:
            people.append(person)

    set_summary_rows(summary_ws, people)
    row_by_person = {summary_ws.cell(row, 1).value: row for row in range(4, 4 + len(people))}

    for (person, group, category), amount in sums.items():
        row = row_by_person.get(person)
        col = summary_columns.get((group, category))
        if row and col:
            summary_ws.cell(row, col).value = money_float(amount)

    for (person, group), amount in totals.items():
        row = row_by_person.get(person)
        col = total_columns.get(group)
        if row and col:
            summary_ws.cell(row, col).value = money_float(amount)

    for row in range(4, 4 + len(people)):
        for col in range(2, summary_ws.max_column + 1):
            summary_ws.cell(row, col).number_format = '#,##0.00'

    for person in template_people:
        if person not in people_seen:
            append_note(notes_ws, "人员不匹配", "Sheet2 原模板有此姓名，但 Sheet1 没有对应交易", person, "", "", Decimal("0"), "")

    for person in sorted(people_seen):
        if person not in template_people:
            append_note(notes_ws, "人员不匹配", "Sheet1 有此姓名，但 Sheet2 原模板没有；已自动新增到 Sheet2", person, "", "", totals_for_person(totals, person), "")

    for (person, group, category), amount in sorted(unsupported.items()):
        append_note(
            notes_ws,
            "模板缺列",
            "此分类在 Sheet2 没有独立列；金额已计入对应通道的总交易额",
            person,
            group,
            category,
            amount,
            "",
        )

    if skipped_rows:
        skipped_total = sum((amount for _, amount in skipped_rows), Decimal("0"))
        locations = ", ".join(f"Sheet1!{row_no}" for row_no, _ in skipped_rows[:10])
        if len(skipped_rows) > 10:
            locations += f" 等 {len(skipped_rows)} 行"
        append_note(notes_ws, "跳过", "缺少姓名/通道/费率，未参与汇总；通常是手工合计行", "", "", "", skipped_total, locations)

    append_note(notes_ws, "汇总说明", f"参与汇总的交易成功明细行数: {detail_count}", "", "", "", sum(totals.values(), Decimal("0")), "")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def totals_for_person(totals: dict, person: str) -> Decimal:
    return sum((amount for (p, _), amount in totals.items() if p == person), Decimal("0"))


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_汇总.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser(description="把 Sheet1 交易明细汇总写入 Sheet2，并生成问题备注。")
    parser.add_argument("input", nargs="?", default="/Users/apple/Desktop/2025.04交易明细.xlsx", help="输入 Excel 文件路径")
    parser.add_argument("-o", "--output", help="输出 Excel 文件路径，默认在原文件旁生成 *_汇总.xlsx")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else default_output_path(input_path)
    build_summary(input_path, output_path)
    print(output_path)


if __name__ == "__main__":
    main()
