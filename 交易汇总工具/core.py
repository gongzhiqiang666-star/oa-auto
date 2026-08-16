#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""交易汇总工具 —— 核心逻辑（计算 + 模板生成）。

计算逻辑与 trade_summary.py 保持一致：
  - 读取 Sheet1（交易明细），按 Sheet2（汇总模板）的结构汇总；
  - 只统计交易状态为“交易成功”的行；
  - 通道按关键字归类：联动=力pos交易、鲲鹏=鲲鹏交易、合利宝=合利宝、星驿付=星驿付；
  - 普通交易按“商户费率”汇总，特惠交易按“支付类型”汇总；
  - 生成“问题备注”“算法说明”工作表。
"""

from __future__ import annotations

from collections import defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE_SHEET = "Sheet1"
SUMMARY_SHEET = "Sheet2"
NOTES_SHEET = "问题备注"
ALGORITHM_SHEET = "算法说明"
GUIDE_SHEET = "使用说明"

REQUIRED_HEADERS = {
    "所属二级",
    "所属通道",
    "商户费率",
    "支付类型",
    "交易金额",
    "交易状态",
}

CHANNEL_KEYWORDS = [
    ("联动", "力pos交易"),
    ("鲲鹏", "鲲鹏交易"),
    ("合利宝", "合利宝"),
    ("星驿付", "星驿付"),
]

# 模板中 Sheet2 的通道布局：通道名 -> 费率/特惠列（最后自动补“总交易额”）
CHANNEL_LAYOUT = [
    ("力pos交易", [
        "0.6交易额", "0.58交易额", "0.57交易额", "0.56交易额", "0.55交易额",
        "0.54交易额", "0.53交易额", "0.38交易额",
        "特惠贷记卡", "特惠GF贷记卡", "特惠PA贷记卡", "特惠MS贷记卡",
    ]),
    ("鲲鹏交易", [
        "0.6交易额", "0.58交易额", "0.57交易额", "0.56交易额", "0.55交易额",
        "0.54交易额", "0.53交易额", "0.48交易额", "0.45交易额", "0.38交易额",
        "0.35交易额", "0.31交易额", "0.3交易额",
    ]),
    ("合利宝", [
        "0.6交易额", "0.57交易额", "0.56交易额", "0.55交易额", "0.38交易额",
    ]),
    ("星驿付", [
        "0.6交易额", "0.58交易额", "0.56交易额", "0.38交易额",
    ]),
]

TEMPLATE_PEOPLE = ["卢红良", "天津张强", "穆延胜"]

DETAIL_HEADERS = [
    "商户名称", "商户编号", "所属代理商", "订单号", "支付类型", "机具sn",
    "卡号", "交易金额", "手续费", "商户费率", "交易时间", "流量卡扣费金额",
    "交易状态", "推送时间", "订单类型", "所属二级", "所属通道", "政策",
]

# 示例数据行：交易状态不是“交易成功”，即使忘记删除也不会参与汇总
EXAMPLE_ROW = [
    "示例商户（请删除本行）", "848100000000001", "A202500001",
    "202604300000000000000000000000000001", "贷记卡", "00000000000000000001",
    "622600******0000", 10000.00, "60.00", "0.6000", "2026-04-30 23:59:59",
    "0.00", "示例（计算前请删除）", "2026-04-30 23:59:59", "普通订单",
    "示例人员", "鲲鹏基础版", "2025基础版-无押金",
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
EXAMPLE_FONT = Font(color="FF0000")
BASE_FONT = "宋体"


# ---------------------------------------------------------------------------
# 计算逻辑（与 trade_summary.py 一致）
# ---------------------------------------------------------------------------

def money(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def money_float(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def rate_header(rate) -> str:
    try:
        normalized = Decimal(str(rate)).normalize()
    except (InvalidOperation, TypeError):
        return f"{rate}交易额"
    return f"{normalized}交易额"


def channel_group(channel) -> str | None:
    if not channel:
        return None
    text = str(channel)
    for keyword, group in CHANNEL_KEYWORDS:
        if keyword in text:
            return group
    return text


def effective_header(pay_type, rate) -> str:
    if pay_type and str(pay_type).startswith("特惠"):
        return str(pay_type)
    return rate_header(rate)


def merged_title(ws, col: int) -> str | None:
    value = ws.cell(1, col).value
    if value:
        return str(value)
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= 1 <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            merged_value = ws.cell(cell_range.min_row, cell_range.min_col).value
            return str(merged_value) if merged_value else None
    return None


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def set_summary_rows(ws, people: list[str]) -> None:
    start_row = 4
    template_row = start_row
    current_rows = max(ws.max_row - start_row + 1, 0)
    needed_rows = len(people)

    if needed_rows > current_rows:
        ws.insert_rows(start_row + current_rows, needed_rows - current_rows)
    elif current_rows > needed_rows:
        ws.delete_rows(start_row + needed_rows, current_rows - needed_rows)

    for offset, person in enumerate(people):
        row = start_row + offset
        copy_row_style(ws, template_row, row)
        ws.cell(row, 1).value = person
        for col in range(2, ws.max_column + 1):
            ws.cell(row, col).value = 0


def clear_notes_sheet(wb):
    if NOTES_SHEET in wb.sheetnames:
        del wb[NOTES_SHEET]
    ws = wb.create_sheet(NOTES_SHEET)
    headers = ["类型", "说明", "姓名", "通道", "分类", "金额", "原始位置/数量"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    widths = [16, 48, 16, 16, 18, 14, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    return ws


def clear_algorithm_sheet(wb):
    if ALGORITHM_SHEET in wb.sheetnames:
        del wb[ALGORITHM_SHEET]
    ws = wb.create_sheet(ALGORITHM_SHEET)
    rows = [
        ("步骤", "说明"),
        ("1", "读取 Sheet1，每一行作为一笔交易明细。"),
        ("2", "只统计交易状态为“交易成功”，且有姓名、通道、费率的明细。"),
        ("3", "姓名取 Sheet1 的“所属二级”，金额取“交易金额”。"),
        ("4", "通道按关键字归类：联动=力pos交易，鲲鹏=鲲鹏交易，合利宝=合利宝，星驿付=星驿付。"),
        ("5", "普通交易按“商户费率”汇总，例如 0.6000 写入 0.6交易额。"),
        ("6", "支付类型以“特惠”开头的交易，按支付类型汇总到对应特惠列。"),
        ("7", "同一姓名、同一通道、同一费率/特惠类型的交易金额相加后写入 Sheet2。"),
        ("8", "每个通道的“总交易额”统计该姓名在该通道下的全部交易金额，即使模板没有独立费率列也计入总额。"),
        ("9", "模板缺列、人员不匹配、无法识别的手工合计行等问题，写入“问题备注”。"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 110
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def append_note(ws, kind, message, person="", channel="", category="", amount=None, location=""):
    amount_value = money_float(amount) if isinstance(amount, Decimal) else amount
    ws.append([kind, message, person, channel, category, amount_value, location])


def totals_for_person(totals: dict, person: str) -> Decimal:
    return sum((amount for (p, _), amount in totals.items() if p == person), Decimal("0"))


def build_summary(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    if SOURCE_SHEET not in wb.sheetnames or SUMMARY_SHEET not in wb.sheetnames:
        raise RuntimeError(f"工作簿必须包含 {SOURCE_SHEET} 和 {SUMMARY_SHEET}")

    detail_ws = wb[SOURCE_SHEET]
    summary_ws = wb[SUMMARY_SHEET]
    notes_ws = clear_notes_sheet(wb)
    clear_algorithm_sheet(wb)

    headers = [cell.value for cell in detail_ws[1]]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise RuntimeError("Sheet1 缺少必要表头: " + ", ".join(sorted(missing)))
    idx = {header: pos for pos, header in enumerate(headers)}

    summary_columns = {}
    total_columns = {}
    for col in range(2, summary_ws.max_column + 1):
        group = merged_title(summary_ws, col)
        header = summary_ws.cell(3, col).value
        if not group or not header:
            continue
        if header == "总交易额":
            total_columns[group] = col
        else:
            summary_columns[(group, str(header))] = col

    template_people = []
    for row in range(4, summary_ws.max_row + 1):
        person = summary_ws.cell(row, 1).value
        if person:
            template_people.append(str(person))

    sums = defaultdict(Decimal)
    totals = defaultdict(Decimal)
    unsupported = defaultdict(Decimal)
    people_seen = set()
    skipped_rows = []
    detail_count = 0

    for row_no, row in enumerate(detail_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        person = row[idx["所属二级"]]
        channel = row[idx["所属通道"]]
        rate = row[idx["商户费率"]]
        pay_type = row[idx["支付类型"]]
        amount = money(row[idx["交易金额"]])
        status = row[idx["交易状态"]]

        if not person or not channel or not rate:
            skipped_rows.append((row_no, amount))
            continue
        if status != "交易成功":
            append_note(notes_ws, "跳过", "交易状态不是交易成功，未参与汇总", person, channel, "", amount, f"Sheet1!{row_no}")
            continue

        detail_count += 1
        person = str(person)
        group = channel_group(channel)
        category = effective_header(pay_type, rate)
        people_seen.add(person)
        totals[(person, group)] += amount

        if (group, category) in summary_columns:
            sums[(person, group, category)] += amount
        else:
            unsupported[(person, group, category)] += amount

    people = []
    for person in template_people:
        if person not in people:
            people.append(person)
    for person in sorted(people_seen):
        if person not in people:
            people.append(person)

    set_summary_rows(summary_ws, people)
    row_by_person = {summary_ws.cell(row, 1).value: row for row in range(4, 4 + len(people))}

    for (person, group, category), amount in sums.items():
        row = row_by_person.get(person)
        col = summary_columns.get((group, category))
        if row and col:
            summary_ws.cell(row, col).value = money_float(amount)

    for (person, group), amount in totals.items():
        row = row_by_person.get(person)
        col = total_columns.get(group)
        if row and col:
            summary_ws.cell(row, col).value = money_float(amount)

    for row in range(4, 4 + len(people)):
        for col in range(2, summary_ws.max_column + 1):
            summary_ws.cell(row, col).number_format = '#,##0.00'

    for person in template_people:
        if person not in people_seen:
            append_note(notes_ws, "人员不匹配", "Sheet2 原模板有此姓名，但 Sheet1 没有对应交易", person, "", "", Decimal("0"), "")

    for person in sorted(people_seen):
        if person not in template_people:
            append_note(notes_ws, "人员不匹配", "Sheet1 有此姓名，但 Sheet2 原模板没有；已自动新增到 Sheet2", person, "", "", totals_for_person(totals, person), "")

    for (person, group, category), amount in sorted(unsupported.items()):
        append_note(
            notes_ws,
            "模板缺列",
            "此分类在 Sheet2 没有独立列；金额已计入对应通道的总交易额",
            person,
            group,
            category,
            amount,
            "",
        )

    if skipped_rows:
        skipped_total = sum((amount for _, amount in skipped_rows), Decimal("0"))
        locations = ", ".join(f"Sheet1!{row_no}" for row_no, _ in skipped_rows[:10])
        if len(skipped_rows) > 10:
            locations += f" 等 {len(skipped_rows)} 行"
        append_note(notes_ws, "跳过", "缺少姓名/通道/费率，未参与汇总；通常是手工合计行", "", "", "", skipped_total, locations)

    append_note(notes_ws, "汇总说明", f"参与汇总的交易成功明细行数: {detail_count}", "", "", "", sum(totals.values(), Decimal("0")), "")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_汇总.xlsx")


# ---------------------------------------------------------------------------
# 模板生成
# ---------------------------------------------------------------------------

GUIDE_ROWS = [
    ("交易汇总工具 · 模板使用说明", ""),
    ("一、填写交易明细（Sheet1）", ""),
    ("1. 打开“交易明细”工作表，把当月的交易明细粘贴进来。", "第一行表头必须保留（列的顺序可以不同）。"),
    ("2. 红色示例行不会参与汇总（交易状态不是“交易成功”），正式计算前请删除或改为真实数据。", "交易状态必须是“交易成功”才会被统计。"),
    ("3. 姓名在“所属二级”列，金额在“交易金额”列。", "商户费率请填写小数，如 0.6000，会汇总到“0.6交易额”。"),
    ("二、汇总模板（Sheet2）", ""),
    ("1. 通道自动归类：联动=力pos交易、鲲鹏=鲲鹏交易、合利宝=合利宝、星驿付=星驿付。", "支付类型以“特惠”开头的交易，按支付类型计入对应特惠列。"),
    ("2. 姓名可以自由增删改。", "Sheet1 中出现的新姓名会自动追加到 Sheet2，并写入“问题备注”。"),
    ("3. 模板没有独立列的分类，金额仍会计入对应通道的“总交易额”。", "所有问题都会写入“问题备注”，请留意查看。"),
    ("三、计算与结果", ""),
    ("1. 保存模板后，打开“交易汇总工具”，点击“导入模板文件”选择这个文件。", "然后点击“开始计算”。"),
    ("2. 结果会自动保存为“原文件名_汇总.xlsx”（与原文件在同一文件夹）。", "可点击“打开文件夹”或“另存为”下载/保存结果。"),
    ("3. 结果包含四个工作表：交易明细、汇总、问题备注、算法说明。", ""),
]


def build_template() -> Workbook:
    """生成一份全新的模板工作簿（结构与原 2025.04 模板一致）。"""
    wb = Workbook()

    # ---------- Sheet1：交易明细 ----------
    ws1 = wb.active
    ws1.title = SOURCE_SHEET
    ws1.append(DETAIL_HEADERS)
    for cell in ws1[1]:
        cell.font = Font(name=BASE_FONT, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.append(EXAMPLE_ROW)
    for cell in ws1[2]:
        cell.font = EXAMPLE_FONT
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
    widths = [26, 16, 13, 30, 10, 16, 16, 12, 10, 10, 20, 12, 12, 20, 10, 12, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---------- Sheet2：汇总模板 ----------
    ws2 = wb.create_sheet(SUMMARY_SHEET)
    col = 2
    for group, sub_headers in CHANNEL_LAYOUT:
        start = col
        ws2.cell(1, col).value = group
        for header in sub_headers:
            ws2.cell(3, col).value = header
            col += 1
        ws2.cell(3, col).value = "总交易额"
        col += 1
        end = col - 1
        ws2.merge_cells(start_row=1, start_column=start, end_row=2, end_column=end)

    ws2.cell(3, 1).value = "姓名"
    for r in range(1, 4):
        for c in range(1, ws2.max_column + 1):
            cell = ws2.cell(r, c)
            cell.font = Font(name=BASE_FONT, bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 42

    for offset, person in enumerate(TEMPLATE_PEOPLE):
        row = 4 + offset
        ws2.cell(row, 1).value = person
        for c in range(2, ws2.max_column + 1):
            ws2.cell(row, c).number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 12
    for c in range(2, ws2.max_column + 1):
        letter = get_column_letter(c)
        header = ws2.cell(3, c).value or ""
        if header == "总交易额":
            ws2.column_dimensions[letter].width = 14
        elif header.startswith("特惠"):
            ws2.column_dimensions[letter].width = 13
        else:
            ws2.column_dimensions[letter].width = 12.5

    # ---------- 使用说明 ----------
    ws3 = wb.create_sheet(GUIDE_SHEET)
    for text, sub in GUIDE_ROWS:
        if text.startswith("交易汇总工具"):
            ws3.append([text])
            ws3.cell(ws3.max_row, 1).font = Font(name=BASE_FONT, size=14, bold=True)
        elif text.endswith("）"):
            ws3.append([text])
            ws3.cell(ws3.max_row, 1).font = Font(name=BASE_FONT, bold=True)
        else:
            ws3.append([text, sub])
    ws3.column_dimensions["A"].width = 70
    ws3.column_dimensions["B"].width = 60
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def save_template(path: Path) -> Path:
    """把模板保存到指定路径。"""
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    build_template().save(path)
    return path


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def calculate(input_path, output_path=None, log=None):
    """执行计算。log 为可选回调函数 log(str)。返回输出文件路径。"""
    def emit(msg):
        if log:
            log(msg)

    input_path = Path(input_path).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"文件不存在：{input_path}")
    output_path = Path(output_path).expanduser().resolve() if output_path else default_output_path(input_path)
    emit(f"正在读取：{input_path.name} ...")
    build_summary(input_path, output_path)
    emit(f"计算完成，结果已保存：{output_path}")
    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        # 命令行模式：python core.py 输入文件 [输出文件]
        out = calculate(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None, print)
        print(out)
    else:
        # 演示：生成模板
        demo = Path(__file__).with_name("交易汇总模板.xlsx")
        save_template(demo)
        print("模板已生成：", demo)
